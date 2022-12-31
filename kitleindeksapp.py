from tkinter import *
from tkinter import messagebox
import tkinter as tk
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

def giris_silme():
    yas_tf.delete(0,'end')
    boy_tf.delete(0,'end')
    agirlik_tf.delete(0,'end')

def vki_hesaplama():
    kg = int(agirlik_tf.get())
    m = int(boy_tf.get())/100
    vki = kg/(m*m)
    vki = round(vki, 1)
    vki_index(vki)

def vki_index(vki):
    
    if vki < 18.5:
        messagebox.showinfo('vucud_kitle_indeksi_hesaplama', f'vki = {vki} Kilonuz zayif.')
        root = tk.Tk()
        root.title('ZayifProgram')
        file ="zayifkiloalma.xlsx"
        wb = load_workbook(file,data_only=TRUE)
        ws= wb.active

        r=0
        for row in ws:
            c = 0 
            for cell in row:
                tk.Label(root,text=cell.value).grid(row=r,column=c)
                c+=1
            r+=1
        root.mainloop()        
    elif (vki > 18.5) and (vki < 24.9):
        messagebox.showinfo('vucud_kitle_indeksi_hesaplama', f'vki = {vki} Kilonuz normal.')
        root = tk.Tk()
        root.title('NormalProgram')
        file ="zayifkiloalma.xlsx"
        wb = load_workbook(file,data_only=TRUE)
        ws= wb.active

        r=0
        for row in ws:
            c = 0 
            for cell in row:
                tk.Label(root,text=cell.value).grid(row=r,column=c)
                c+=1
            r+=1
        root.mainloop()         
    elif (vki > 24.9) and (vki < 29.9):
        messagebox.showinfo('vucud_kitle_indeksi_hesaplama', f'vki = {vki} Fazla kilolusunuz.')
        root = tk.Tk()
        root.title('Kilolu Program')
        file ="zayifkiloalma.xlsx"
        wb = load_workbook(file,data_only=TRUE)
        ws= wb.active

        r=0
        for row in ws:
            c = 0 
            for cell in row:
                tk.Label(root,text=cell.value).grid(row=r,column=c)
                c+=1
            r+=1
        root.mainloop()         
    elif (vki > 29.9):
        messagebox.showinfo('vucud_kitle_indeksi_hesaplama', f'vki = {vki} Obezsiniz.\nSağliğiniz için doktor kontrolü altinda bir diyet programi aliniz.')
   
    else:
        messagebox.showerror('vucud_kitle_indeksi_hesaplama', 'Bir hata oluştu tekrar deneyiniz!')   

ws = Tk()
ws.title('PythonProje')
ws.geometry('400x300')
ws.config(bg='#686e70')

var = IntVar()

frame = Frame(
    ws,
    padx=10, 
    pady=10
)
frame.pack(expand=True)


age_lb = Label(
    frame,
    text="Yaş Giriniz"
)
age_lb.grid(row=1, column=1)

yas_tf = Entry(
    frame, 
)
yas_tf.grid(row=1, column=2, pady=5)

gen_lb = Label(
    frame,
    text='Cinsiyet seçiniz'
)
gen_lb.grid(row=2, column=1)

frame2 = Frame(
    frame
)
frame2.grid(row=2, column=2, pady=5)

male_rb = Radiobutton(
    frame2,
    text = 'Erkek',
    variable = var,
    value = 1
)
male_rb.pack(side=LEFT)

female_rb = Radiobutton(
    frame2,
    text = 'Kadin',
    variable = var,
    value = 2
)
female_rb.pack(side=RIGHT)

height_lb = Label(
    frame,
    text="Boyunuzu giriniz. (cm)  "
)
height_lb.grid(row=3, column=1)

weight_lb = Label(
    frame,
    text="Kilonuzu giriniz (kg)  ",

)
weight_lb.grid(row=4, column=1)

boy_tf = Entry(
    frame,
)
boy_tf.grid(row=3, column=2, pady=5)

agirlik_tf = Entry(
    frame,
)
agirlik_tf.grid(row=4, column=2, pady=5)

frame3 = Frame(
    frame
)
frame3.grid(row=5, columnspan=3, pady=10)

cal_btn = Button(
    frame3,
    text='Hesapla',
    command=vki_hesaplama
)
cal_btn.pack(side=LEFT)

reset_btn = Button(
    frame3,
    text='Sil',
    command=giris_silme
)
reset_btn.pack(side=LEFT)

exit_btn = Button(
    frame3,
    text='Kapat',
    command=lambda:ws.destroy()
)
exit_btn.pack(side=RIGHT)

ws.mainloop()